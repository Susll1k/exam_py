
import psycopg2
from psycopg2 import OperationalError
from tabulate import tabulate
from openpyxl import Workbook


while True:
    print(f"Здраствуйте, в нашем магазине вы можете приобрести товары электроники.\nЧто вы хотите?")
    task=int(input('1)Добавление в корзину\n2)Просмотр списка товаров.\n3)Удаление записи из базы данных.\n4)Оплата содержимого корзины.\n:'))
    if task >=1 and task <= 4:
        break



def connect_to_db():
    try:
        conn = psycopg2.connect(host='127.0.0.1', port=5432, dbname='store', user='postgres', password='postgres')
        return conn
    except OperationalError as e:
        print(e)

connection = connect_to_db()
cursor = connection.cursor()
querystart = f'select * from store'
cursor.execute(querystart)
response = cursor.fetchall()
my_dict = []

for row in response:
    my_dict.append([row[0],row[1],row[2],row[3],row[4]])

def add():

    cursor = connection.cursor()
    querystart = f'select * from store'
    cursor.execute(querystart)
    response = cursor.fetchall()

    id= int(input('Id товара который вы хотите добавить в корзину: '))
    while True:
        quantly= int(input('Количество?: '))
        if quantly < response[id][4]:
            cursor.execute(f"INSERT INTO store (availability) VALUES ({response[id][4]-quantly})")
            connection.commit()
            break

    

    query = f"INSERT INTO cart (name, price, quantiy) VALUES ('{response[id][1]}', '{response[id][3]}', {quantly})"
    cursor.execute(query)
    connection.commit()

def delete():
    id= int(input('Введи id: '))

    cursor = connection.cursor()
    query = f'Delete from cart where id={id}'
    cursor.execute(query)
    connection.commit()

def show_table():

    cursor = connection.cursor()
    querystart = f'select * from store'
    cursor.execute(querystart)


    response = cursor.fetchall()



    print(tabulate(my_dict, headers=['id','name','description','price','availability'], tablefmt="pretty"))


def cheak():
    wb = Workbook()


    ws = wb.active

    cursor = connection.cursor()
    querystart = f'select * from cart'
    cursor.execute(querystart)


    response = cursor.fetchall()

    ws['A1'] = "№"
    ws['B1'] = "Name"
    ws['C1'] = "Description"
    ws['D1'] = "Price"

    price=0

    for i in range(len(response)):
        ws[f'A{i+2}'] = response[i][0]
        ws[f'B{i+2}'] = response[i][1]
        ws[f'C{i+2}'] = response[i][2]
        ws[f'D{i+2}'] = response[i][3]
        price = price + response[i][2]
    ws['E1'] = "Total price"
    ws['E2'] = price


    wb.save("cheak.xlsx")

if task == 1:
    add()
elif task == 2:
    show_table()
elif task == 3:
    delete()
elif task == 4:
    cheak()

    



    

